#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
GBFR 伤害字体 mod 编辑器 (tkinter, 单文件 GUI)

功能:
  - 列出全部 112 精灵(按后缀组别可折叠分组): ①元素名 ②坐标 ③原生图(真实 UV 比例)
  - ④每精灵 三选一: 保留 / 屏蔽 / 替换  (列头: 全部保留/屏蔽/替换)
  - ⑤每精灵: 导入 PNG + 占比(无上限) + 替换预览
              (列头: 批量导入文件夹 + 批量占比)
  - 控制台日志
  - 字体(TTF/OTF) -> 批量生成上色 PNG (每组自定义颜色 + 加粗/斜体)
  - 生成 Reloaded-II mod zip

运行期依赖: numpy, PIL, tkinter, resources/(原生图集+头+texconv)
"""
import os, sys, json, threading, datetime, re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
import numpy as np
from PIL import Image, ImageTk, ImageDraw
import io
import openpyxl
from openpyxl.drawing.image import Image as XLImage

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import core

# 列宽 (①名缩窄 ②坐标缩窄 ⑤替换图加宽放预览+控件)
COL_W = [150, 90, 150, 250, 300]
COL_NAME, COL_COORD, COL_NATIVE, COL_ACTION, COL_REPLACE = range(5)
NATIVE_MAX = (140, 92)
PREVIEW_MAX = (140, 92)

# 伤害数字大小(hud_param.json damage_ 段) 的 UI 显示名 -> i18n key
SIZE_LABEL_KEYS = {
    "spArtsLinkAttackSize_": "sz_size",
    "height_": "sz_height",
    "length_": "sz_length",
    "viewTime_": "sz_viewtime",
    "space_": "sz_space",
    "healSpace_": "sz_healspace",
    "commonSize_": "sz_commonsize",
    "criticalSize_": "sz_criticalsize",
    "playerSize_": "sz_playersize",
    "playerCriticalSize_": "sz_playercriticalsize",
}
SIZE_HINT_KEYS = {
    "spArtsLinkAttackSize_": "sz_size_hint",
    "height_": "sz_height_hint",
    "length_": "sz_length_hint",
    "viewTime_": "sz_viewtime_hint",
    "space_": "sz_space_hint",
    "healSpace_": "sz_healspace_hint",
    "commonSize_": "sz_commonsize_hint",
    "criticalSize_": "sz_criticalsize_hint",
    "playerSize_": "sz_playersize_hint",
    "playerCriticalSize_": "sz_playercriticalsize_hint",
}
# 组别显示名 -> i18n key (member / normal 无映射, 保持原名)
GROUP_TR = {
    "add": "grp_add", "damage": "grp_damage", "effect": "grp_effect",
    "guard01": "grp_guard01", "guard02": "grp_guard02", "heal": "grp_heal",
    "link": "grp_link",
}
# 伤害数字控制默认值 (hud_param.json damage_ 段的原生基准):
#   spArtsLinkAttackSize_=1.0(大小) / height_=1.0(垂直位置) / length_=80.0(水平位置) / space_=-12.0(间距)
DEFAULT_SIZE_SCALE = {
    "spArtsLinkAttackSize_": 1.0,
    "height_": 1.0,
    "length_": 80.0,
    "viewTime_": 0.2,
    "space_": -12.0,
    "healSpace_": -32.0,
    "commonSize_": 0.42,
    "criticalSize_": 0.33,
    "playerSize_": 1.0,
    "playerCriticalSize_": 0.8,
}

# 队友伤害颜色 (damage_ 段) 原生基准; 工具内用 0-255 RGBA, 写入时 /255
COLOR_KEYS = ["normalAttackColor_", "spAttackColor_"]
DEFAULT_COLOR = {
    "normalAttackColor_": (210, 235, 240, 128),
    "spAttackColor_": (240, 235, 180, 128),
}
COLOR_LABEL_KEYS = {
    "normalAttackColor_": "sz_col_normal",
    "spAttackColor_": "sz_col_sp",
}
COLOR_HINT_KEYS = {
    "normalAttackColor_": "sz_col_normal_hint",
    "spAttackColor_": "sz_col_sp_hint",
}

# 浮动系数 (damage_ 段 rangeNear_/rangeFar_)
# 原生基准: rangeNear_=[100,300,350,0]  rangeFar_=[100,200,250,0]
# 用户只设一个 alpha 系数, 构建时逐元素相乘: result[i] = native[i] * alpha
DEFAULT_RANGE = {
    "rangeNear_": [100, 300, 350, 0],
    "rangeFar_":  [100, 200, 250, 0],
}
DEFAULT_RANGE_ALPHA = 1.0


# ───────────────── 国际化 (i18n) ─────────────────
LANG = "zh"   # "zh"(默认) 或 "en"

I18N = {
    "zh": {
        "app_title": "GBFR 伤害字体 Mod 编辑器  v40   UP: bilibili / Dangoooooo  QQ:1041271418",
        "gen_mod": "▶ 生成 Mod",
        "ttf_tab": "用字体生成png",
        "batch_replace": "用png图片批量替换全组别的个别元素",
        "size_control": "伤害大小/浮动/队友伤害颜色",
        "mod_name": "Mod 名称:",
        "mod_out": "Mod 输出目录:",
        "browse": "浏览",
        "col_name": "①元素名",
        "col_coord": "②坐标",
        "col_native": "③原生图 (真实比例)",
        "col_action": "④动作",
        "col_replace": "⑤替换图片",
        "all_keep": "全部保留",
        "all_block": "全部屏蔽",
        "all_replace": "全部替换",
        "batch_import": "批量从文件夹导入...",
        "ratio": "占比",
        "stretch": "铺满",
        "console": "控制台",
        "grp_add": "底层附加组",
        "grp_damage": "受伤组",
        "grp_effect": "特效组",
        "grp_guard01": "烂肉质组1",
        "grp_guard02": "烂肉质组2",
        "grp_heal": "回血组",
        "grp_link": "属性克制组",
        "grp_elems": "· %d 个元素",
        "grp_replaceable": " (可替换 %d)",
        "act_keep": "保留",
        "act_block": "屏蔽",
        "act_replace": "替换",
        "after_replace": "（替换后显示）",
        "import": "导入..",
        "ratio_lbl": "占比",
        "stretch_lbl": "铺满",
        "sz_win_title": "伤害大小 / 浮动 / 队友伤害颜色 (hud_param.json)",
        "sz_desc": "伤害数字外观与队友颜色 (写入 mod 的 hud_param.json → damage_ 段)。仅 spArtsLinkAttackSize_ 实际改变数字大小，其佗控制位置/间距/时间；队友颜色用 RGBA(0-255)。回车或失焦应用，重置恢复原生基准。",
        "sz_size": "属性克制时变化系数",
        "sz_height": "垂直位置",
        "sz_length": "水平位置",
        "sz_space": "数字间距",
        "sz_size_hint": "建议 0.3~2；0=关闭，1.0=原生",
        "sz_height_hint": "推荐-2~2；0=伤害字幕区的水平轴线过屏幕中心",
        "sz_length_hint": "推荐-200~200；0=伤害字幕区的垂直轴线过屏幕中心",
        "sz_space_hint": "推荐 -40~40",
        "sz_close": "关闭",        "sz_note": "回车或失去焦点即应用; 重置恢复原生基准值。",        "sz_reset": "重置",        "ttf_win_title": "字体生成选项",
        "sz_viewtime": "显示时间(秒)",
        "sz_healspace": "治愈间距",
        "sz_commonsize": "队友普通大小",
        "sz_criticalsize": "队友暴击大小",
        "sz_playersize": "玩家普通大小",
        "sz_playercriticalsize": "玩家暴击大小",
        "sz_viewtime_hint": "秒; 推荐0.2",
        "sz_healspace_hint": "推荐 -40~40",
        "sz_commonsize_hint": "推荐 0.2~1",
        "sz_criticalsize_hint": "推荐 0.2~1",
        "sz_playersize_hint": "推荐 0.3~1.2",
        "sz_playercriticalsize_hint": "推荐 0.3~1.2",
        "sz_col_normal": "队友普通颜色",
        "sz_col_sp": "队友克制颜色",
        "sz_col_normal_hint": "RGBA; 0-255, A=不透明度",
        "sz_col_sp_hint": "RGBA; 0-255, A=不透明度",
        "sz_sec_pos": "位置与时间",
        "sz_sec_space": "间距",
        "sz_sec_size": "大小",
        "sz_sec_color": "队友伤害颜色 (RGBA)",
        "sz_sec_float": "浮动 (Spread)",
        "sz_range_alpha": "浮动系数 alpha",
        "sz_range_alpha_hint": "1.0=原生(不变); 越大数字分布越散, 越小越紧凑; 建议 0.5~3.0",
        "col_pick": "选择 %s 颜色",
        "ttf_heading": "字体上色 / 样式设置",
        "ttf_file": "文件",
        "ttf_font": "字体文件:",
        "ttf_out": "输出目录:",
        "ttf_groups": "各组颜色与样式 (点击色块自定义颜色)",
        "ttf_items": "  共%d条",
        "ttf_opacity": "不透明度:",
        "ttf_bold": "加粗",
        "ttf_italic": "斜体",
        "ttf_reset": "恢复默认",
        "ttf_gen": "▶ 生成 PNG",
        "ttf_tip": "提示: 默认颜色=各组原生主色均值。",
        "ttf_pick_font": "选择字体文件",
        "ttf_pick_out": "选择生成的 PNG 输出文件夹",
        "ttf_pick_color": "选择 %s 组颜色",
        "ttf_err_font": "错误",
        "ttf_err_font_msg": "请先选择有效的字体文件(TTF/OTF)。",
        "ttf_err_out": "错误",
        "ttf_err_out_msg": "请选择 PNG 输出目录。",
        "ttf_gen_log": "字体生成中: %s",
        "ttf_done_log": "字体已生成 %d 张并自动设为替换。",
        "ttf_done_title": "字体生成完成",
        "ttf_done_msg": "已生成 %d 张上色 PNG 到:\n%s\n并已自动设为替换(可调占比)。",
        "ttf_fail_log": "字体失败: %s",
        "ttf_fail_title": "字体生成失败",
        "br_win_title": "用png图片批量替换全组别的个别元素 (按编号 00-12)",
        "br_desc": "按编号批量替换: 给每类编号导入一张图, 点'组内同类别覆盖'可一次性替换所有组中该编号的精灵",
        "br_apply": "组内同类别覆盖",
        "br_import": "导入替换图片",
        "br_unimported": "（未导入）",
        "br_import_btn": "导入图片",
        "br_row_frame": "%s   编号 %s    出现于: %s",
        "br_warn": "提示",
        "br_warn_msg": "请先为本类导入一张图片。",
        "br_done": "完成",
        "br_done_msg": "已将图片替换到编号 %s 的所有精灵, 共 %d 处。",
        "br_cover_log": "组内同类别覆盖: 编号 %s <- 已选图片 (%s, 影响 %d 处)",
        "br_scale_mode": "占比 %d%%",
        "br_none": "无",
        "about_title": "关于",
        "about_body": "GBFR 伤害字体 Mod 编辑器  v40\n\n傻瓜式制作 GBFR 伤害数字/符号字体 mod。\n①名 ②坐标 ③原生真实比例图 ④保留/屏蔽/替换 ⑤导入+占比+预览。\n按后缀组别折叠分组; 字体(TTF/OTF)生成支持每组自定义颜色、加粗、斜体。\n支持批量导入文件夹; 支持按编号批量替换全组别元素。\n新增: 伤害大小/浮动/队友伤害颜色 (hud_param.json → damage_): 属性克制系数 spArtsLinkAttackSize_ 0.3~2(0=关闭)、垂直位置 height_(0=轴线过屏中心) 推荐-2~2、水平位置 length_(0=轴线过屏中心) 推荐-200~200、字符间距 space_ 推荐-40~40、治愈间距 healSpace_ 推荐-40~40、显示时间 viewTime_ 推荐0.2s、队友大小 commonSize_/criticalSize_ 推荐0.2~1、玩家大小 playerSize_/playerCriticalSize_ 推荐0.3~1.2、队友颜色 normalAttackColor_/spAttackColor_ (RGBA 0-255)、浮动系数 alpha (乘以原生值, 1.0=原生, 越大越散)。\n\n依赖 Reloaded-II + gbfrelink.utility.manager。",
        "build_out_missing": "错误",
        "build_out_missing_msg": "请先设置输出目录",
        "confirm_title": "确认",
        "confirm_msg": "没有设置任何替换/屏蔽, 将生成与原版相同的 mod。继续?",
        "build_start": "开始生成 Mod (替换=%d, 屏蔽=%d, 保留=%d) ...",
        "build_done_title": "生成完成",
        "build_done_msg": "Mod 已生成:\n%s",
        "build_fail_log": "生成失败: %s",
        "core_build_start": "开始构建 mod: 主图集 + 高分辨率图集",        "core_no_tpl": "  ⚠ 未找到 resources/hud_param.json 模板, 跳过大小控制写入",        "core_hud_fail": "  ⚠ hud_param.json 写入失败: %s",        "core_set_replace_no_img": "  ⚠ %s 设为替换但无图片, 按保留处理",        "core_repl": "  [%s] 替换=%d 屏蔽=%d 保留=%d -> %s",        "core_pack_done": "  ✅ 打包完成: %s (%s bytes)",        "core_ttf_gen": "  ✅ TTF 生成 %d 张 PNG -> %s",        "core_no_texconv": "未找到 texconv.exe (resources 内)",        "build_fail_title": "生成失败",
        "import_fail_title": "导入失败",
        "batch_no_png_title": "批量导入",
        "batch_no_png_msg": "未在文件夹中找到与元素名匹配的 PNG。\n文件名需为 元素名.png (如 hud_nmbtl_damage00.png)",
        "browse_out_title": "选择 Mod 输出目录",
        "browse_batch_title": "选择含 {元素名}.png 的文件夹",
        "log_loaded": "已加载原生图集: 主 %dx%d / 高分辨率 %dx%d, 共 %d 精灵 (可替换 %d)",
        "log_tip": "提示: ③列显示 UV 真实比例; ④选替换后请在⑤导入图片或先用字体生成。",
        "log_bulk": "已批量设为: %s",
        "log_import": "导入 %s <- %s",
        "log_batch": "批量导入 %d 张 (占比 %d%%)",
        "log_batch_skip_effect": "已跳过 effect 组 %d 处 (不参与批量替换)",
        "log_export": "已导出配置: %s",
        "exp_col": "导出…",
        "exp_all": "将全列导出…",
        "exp_names_title": "导出元素名单",
        "exp_coords_title": "导出坐标表",
        "exp_pngs_title": "选择原生 PNG 导出文件夹",
        "exp_all_title": "导出全列表格",
        "log_export_names": "已导出元素名单: %s",
        "log_export_coords": "已导出坐标表: %s",
        "log_export_pngs": "已导出 %d 张原生 PNG 到 %s",
        "log_export_all": "已导出全列表格: %s",
        "png_export_fail": "PNG 导出失败 %s: %s",        "exp_pngs_done": "已导出 %d 张原生 PNG 到:\n%s",
        "exp_done": "已导出:\n%s",
        "log_import_cfg": "已导入配置: %s",
        "lang_toggle": "EN",
        "tutorial": "教程",
    },
    "en": {
        "app_title": "GBFR Damage Font Mod Editor  v40   UP: bilibili / Dangoooooo  QQ:1041271418",
        "gen_mod": "▶ Build Mod",
        "ttf_tab": "Generate PNG from Font",
        "batch_replace": "Batch Replace Elements (by code 00-12)",
        "size_control": "Damage Size / Float / Teammate Color",
        "mod_name": "Mod Name:",
        "mod_out": "Mod Output Dir:",
        "browse": "Browse",
        "col_name": "Element Name",
        "col_coord": "Coord",
        "col_native": "Native (real ratio)",
        "col_action": "Action",
        "col_replace": "Replace Image",
        "all_keep": "All Keep",
        "all_block": "All Block",
        "all_replace": "All Replace",
        "batch_import": "Batch Import from Folder...",
        "ratio": "Scale",
        "stretch": "Stretch",
        "console": "Console",
        "grp_add": "Base Add",
        "grp_damage": "Damage",
        "grp_effect": "Effect",
        "grp_guard01": "Flesh Weak 1",
        "grp_guard02": "Flesh Weak 2",
        "grp_heal": "Heal",
        "grp_link": "Element Counter",
        "grp_elems": "· %d elements",
        "grp_replaceable": " (%d replaceable)",
        "act_keep": "Keep",
        "act_block": "Block",
        "act_replace": "Replace",
        "after_replace": "(preview after replace)",
        "import": "Import..",
        "ratio_lbl": "Scale",
        "stretch_lbl": "Stretch",
        "sz_win_title": "Damage Size / Float / Teammate Color (hud_param.json)",
        "sz_desc": "Damage number appearance & teammate color (writes mod hud_param.json → damage_). Only spArtsLinkAttackSize_ changes size; others control position/spacing/time. Teammate color uses RGBA (0-255). Enter/focus-out applies; Reset restores native base.",
        "sz_size": "Elemental Advantage Scale Factor",
        "sz_height": "Vertical Pos",
        "sz_length": "Horizontal Pos",
        "sz_space": "Spacing",
        "sz_size_hint": "Suggest 0.3~2; 0=off, 1.0=native",
        "sz_height_hint": "Recommend -2~2; 0=damage area horizontal axis through screen center",
        "sz_length_hint": "Recommend -200~200; 0=damage area vertical axis through screen center",
        "sz_space_hint": "Recommend -40~40",
        "sz_close": "Close",        "sz_note": "Enter or focus-out applies; Reset restores native base values.",        "sz_reset": "Reset",        "ttf_win_title": "Font Options",
        "sz_viewtime": "Show Time (s)",
        "sz_healspace": "Heal Spacing",
        "sz_commonsize": "Ally Normal Size",
        "sz_criticalsize": "Ally Crit Size",
        "sz_playersize": "Player Normal Size",
        "sz_playercriticalsize": "Player Crit Size",
        "sz_viewtime_hint": "sec; recommend 0.2",
        "sz_healspace_hint": "recommend -40~40",
        "sz_commonsize_hint": "recommend 0.2~1",
        "sz_criticalsize_hint": "recommend 0.2~1",
        "sz_playersize_hint": "recommend 0.3~1.2",
        "sz_playercriticalsize_hint": "recommend 0.3~1.2",
        "sz_col_normal": "Ally Normal Color",
        "sz_col_sp": "Ally Advantage Color",
        "sz_col_normal_hint": "RGBA; 0-255, A=opacity",
        "sz_col_sp_hint": "RGBA; 0-255, A=opacity",
        "sz_sec_pos": "Position & Time",
        "sz_sec_space": "Spacing",
        "sz_sec_size": "Sizes",
        "sz_sec_color": "Teammate Damage Color (RGBA)",
        "sz_sec_float": "Float / Spread",
        "sz_range_alpha": "Spread factor alpha",
        "sz_range_alpha_hint": "1.0=native; larger=more spread, smaller=more compact; suggest 0.5~3.0",
        "col_pick": "Pick %s color",
        "ttf_heading": "Font Color / Style",
        "ttf_file": "File",
        "ttf_font": "Font File:",
        "ttf_out": "Output Dir:",
        "ttf_groups": "Per-Group Color & Style (click swatch to pick)",
        "ttf_items": "  %d items",
        "ttf_opacity": "Opacity:",
        "ttf_bold": "Bold",
        "ttf_italic": "Italic",
        "ttf_reset": "Reset Default",
        "ttf_gen": "▶ Generate PNG",
        "ttf_tip": "Tip: default color = each group's native average color.",
        "ttf_pick_font": "Select Font File",
        "ttf_pick_out": "Select PNG Output Folder",
        "ttf_pick_color": "Select %s group color",
        "ttf_err_font": "Error",
        "ttf_err_font_msg": "Please select a valid font file (TTF/OTF) first.",
        "ttf_err_out": "Error",
        "ttf_err_out_msg": "Please select a PNG output folder.",
        "ttf_gen_log": "Font generating: %s",
        "ttf_done_log": "Font generated %d images and auto-set as replace.",
        "ttf_done_title": "Font Generation Done",
        "ttf_done_msg": "Generated %d colored PNGs to:\n%s\nand auto-set as replace (scale adjustable).",
        "ttf_fail_log": "Font failed: %s",
        "ttf_fail_title": "Font Generation Failed",
        "br_win_title": "Batch Replace Elements by Code (00-12)",
        "br_desc": "Batch replace by code: import one image per code, click 'Apply to Group' to replace all sprites with that code across groups.",
        "br_apply": "Apply to Group",
        "br_import": "Import Replace Image",
        "br_unimported": "(not imported)",
        "br_import_btn": "Import Image",
        "br_row_frame": "%s   code %s   appears in: %s",
        "br_warn": "Tip",
        "br_warn_msg": "Please import an image for this code first.",
        "br_done": "Done",
        "br_done_msg": "Replaced image for all sprites with code %s, %d entries.",
        "br_cover_log": "Apply to Group: code %s <- selected image (%s, %d entries)",
        "br_scale_mode": "Scale %d%%",
        "br_none": "none",
        "about_title": "About",
        "about_body": "GBFR Damage Font Mod Editor  v40\n\nEasily create GBFR damage number / symbol font mods.\nCols: Name | Coord | Native (real ratio) | Keep/Block/Replace | Import+Scale+Preview.\nCollapsible groups by suffix; Font (TTF/OTF) supports per-group color, bold, italic.\nBatch folder import; batch replace all groups by code.\nNew: Damage Size / Float / Teammate Color control (hud_param.json → damage_): ElemAdv Scale spArtsLinkAttackSize_ 0.3~2(0=off), Vertical height_(0=axis thru center) recommend -2~2, Horizontal length_(0=axis thru center) recommend -200~200, Spacing space_ recommend -40~40, Heal Spacing healSpace_ recommend -40~40, Show viewTime_ recommend 0.2s, Ally sizes commonSize_/criticalSize_ recommend 0.2~1, Player sizes playerSize_/playerCriticalSize_ recommend 0.3~1.2, teammate color normalAttackColor_/spAttackColor_ (RGBA 0-255), float alpha (multiplier on native values, 1.0=native, larger=more spread).\n\nRequires Reloaded-II + gbfrelink.utility.manager.",
        "build_out_missing": "Error",
        "build_out_missing_msg": "Please set the output directory first",
        "confirm_title": "Confirm",
        "confirm_msg": "No replace/block set; will generate an identical mod. Continue?",
        "build_start": "Start building Mod (replace=%d, block=%d, keep=%d) ...",
        "build_done_title": "Build Done",
        "build_done_msg": "Mod generated:\n%s",
        "build_fail_log": "Build failed: %s",
        "core_build_start": "Building mod: main atlas + hi-res atlas",        "core_no_tpl": "  ⚠ resources/hud_param.json template not found, skipping size control write",        "core_hud_fail": "  ⚠ hud_param.json write failed: %s",        "core_set_replace_no_img": "  ⚠ %s set as replace but no image, treated as keep",        "core_repl": "  [%s] replace=%d block=%d keep=%d -> %s",        "core_pack_done": "  ✅ Packed: %s (%s bytes)",        "core_ttf_gen": "  ✅ TTF generated %d PNGs -> %s",        "core_no_texconv": "texconv.exe not found (inside resources)",        "build_fail_title": "Build Failed",
        "import_fail_title": "Import Failed",
        "batch_no_png_title": "Batch Import",
        "batch_no_png_msg": "No PNG matching element names found in folder.\nFile must be element.png (e.g. hud_nmbtl_damage00.png)",
        "browse_out_title": "Select Mod Output Directory",
        "browse_batch_title": "Select folder containing {element}.png",
        "log_loaded": "Loaded native atlas: main %dx%d / hi-res %dx%d, %d sprites (%d replaceable)",
        "log_tip": "Tip: the Native column shows real UV ratio; pick Replace in Action then import in Replace Image, or use a font first.",
        "log_bulk": "Bulk set to: %s",
        "log_import": "Import %s <- %s",
        "log_batch": "Batch imported %d images (scale %d%%)",
        "log_batch_skip_effect": "Skipped effect group %d (excluded from batch replace)",
        "log_export": "Config exported: %s",
        "exp_col": "Export…",
        "exp_all": "Export all cols…",
        "exp_names_title": "Export element names",
        "exp_coords_title": "Export coordinates",
        "exp_pngs_title": "Choose folder for native PNGs",
        "exp_all_title": "Export all columns",
        "log_export_names": "Element names exported: %s",
        "log_export_coords": "Coordinates exported: %s",
        "log_export_pngs": "Exported %d native PNGs to %s",
        "log_export_all": "All-columns table exported: %s",
        "png_export_fail": "PNG export failed %s: %s",        "exp_pngs_done": "Exported %d native PNGs to:\n%s",
        "exp_done": "Exported:\n%s",
        "log_import_cfg": "Config imported: %s",
        "lang_toggle": "中文",
        "tutorial": "Tutorial",
    },
}


TUTORIAL_TEXT = {
    "zh": {
        "title": "教程 · 如何使用本工具",
        "body": (
"一、界面速览\n"
"· 顶部工具栏：生成 Mod ｜ 用字体生成png ｜ 批量替换 ｜ 数字大小控制 ｜ 教程 ｜ 语言切换(EN/中文)\n"
"· 第二行：Mod 名称(可改) ＋ Mod 输出目录(默认 ./mod_output) ＋ 浏览\n"
"· 表格五列：①元素名 ②坐标 ③原生图(真实比例) ④动作(保留/屏蔽/替换) ⑤替换图片(预览＋导入＋占比＋铺满)\n"
"· ⑤列头还带：批量从文件夹导入... ＋ 占比 ＋ 铺满\n"
"· 列头①②③下方各有『导出…』按钮(名单/坐标/原生PNG); 工具栏『将全列导出…』可一键导出整表 Excel\n"
"· 表格按组别可折叠(底层附加/受伤/特效/烂肉质1/烂肉质2/回血/属性克制)，点组标题展开/收起\n"
"· 底部：控制台(操作日志)\n"
"\n"
"二、替换字形(三种方式，可任选混用)\n"
"1) 用字体(TTF/OTF)生成彩色 PNG(最省事)\n"
"   · 点“用字体生成png” -> 设置框\n"
"   · 选字体：在“字体文件”栏点浏览，选你的 .ttf 或 .otf(默认 ./LiXuKeShuFa-1.ttf，可换任意字体)\n"
"   · 字体 PNG 输出目录：设置框里的“输出目录”(默认 ./font_output)，是生成 PNG 的保存文件夹，相对 exe 所在目录(直接双击 exe 则在 exe 旁新建 font_output)。生成后自动导入主界面并设为替换\n"
"   · 每组颜色/不透明度/加粗/斜体可逐组自定义(默认=各组原生主色均值；属性克制组默认金黄)\n"
"   · 点“生成 PNG”开始，完成弹窗告知张数与路径\n"
"2) 单个手动导入：在⑤列点某行“导入..”选 PNG -> 自动设为替换，可调占比/铺满\n"
"3) 批量从文件夹导入图片：在⑤列头点“批量从文件夹导入...”选文件夹，程序按 文件名=元素名(如 hud_nmbtl_damage00.png) 自动匹配导入并设为替换；占比/铺满取自列头右侧输入\n"
"4) 按编号批量替换(00-12)：工具栏“批量替换”弹窗，13 类编号各给一张图，点“组内同类别覆盖”一次替换所有组该编号\n"
"\n"
        "三、调整数字外观/位置/队友颜色(可选)\n"
        "· 点“伤害大小/浮动/队友伤害颜色” -> 多个分组输入框：\n"
        "   [位置与时间] height_: 伤害字幕区总体垂直位置，0时水平轴线过屏幕中心 推荐-2~2；length_: 总体水平位置，0时垂直轴线过屏幕中心 推荐-200~200；viewTime_: 显示时间(秒) 推荐0.2\n"
        "   [间距] space_: 每串伤害数字内部字符拥挤程度 推荐-40~40；healSpace_: 每串治愈数字内部字符拥挤程度 推荐-40~40\n"
        "   [大小] commonSize_(未暴击队友) 推荐0.2~1；criticalSize_(暴击队友) 推荐0.2~1；playerSize_(未暴击玩家) 推荐0.3~1.2；playerCriticalSize_(暴击玩家) 推荐0.3~1.2；spArtsLinkAttackSize_(属性克制系数) 0.3~2(0=关闭,1=原生)\n"
        "   [浮动] 浮动系数 alpha：一个数控制 rangeNear_/rangeFar_ 全部 8 个原生值的整体缩放；1.0=原生不变，越大数字分布越散，越小越紧凑；建议 0.5~3.0\n"
        "   [队友颜色 RGBA] normalAttackColor_(无克制) / spAttackColor_(克制)：点色块开调色板，或填 R/G/B/A(0-255, A=不透明度)\n"
        "· 直接输入浮点，回车或失焦生效；“全部重置”恢复原生基准\n"
"\n"
"四、导出(可选)\n"
"· 列头①②③下方各有『导出…』按钮，弹窗已按当前语言自动填好默认文件名，可直接保存或改名：\n"
"   - 列1(元素名)：导出 Excel『元素名列表.xlsx』(序号＋元素名)\n"
"   - 列2(坐标)：导出 Excel『坐标表.xlsx』(序号＋元素名＋X＋Y＋宽＋高)\n"
"   - 列3(原生图)：选文件夹，把每行原生图另存为『元素名.png』(默认打开“原生图”文件夹)\n"
"· 工具栏『数字大小控制』右侧的『将全列导出…』：导出整表 Excel，直接带缩略图(元素名/坐标/尺寸/原生图/动作/替换图)\n"
"· 中英文界面默认名不同：中文如上；英文为 element_names / coordinates / all_columns / native_pngs\n"
"\n"
"五、生成 Mod\n"
"· 确认第二行 Mod 输出目录(默认 ./mod_output，注意和字体 PNG 输出目录不是同一个)\n"
"· 填 Mod 名称(可选)\n"
"· 点“生成 Mod” -> 生成 zip 并自动打开所在文件夹\n"
"\n"
"六、安装到游戏\n"
"· 需 Reloaded-II ＋ gbfrelink.utility.manager，把生成的 mod 加载启动即可\n"
"\n"
"注意：两个“输出目录”别搞混\n"
"· 顶部 Mod 输出目录 = 最终 mod 压缩包(默认 ./mod_output)\n"
"· “用字体生成png”对话框里的输出目录 = 字体彩色 PNG(默认 ./font_output，相对 exe 目录)\n"
        ),
    },
    "en": {
        "title": "Tutorial · How to use this tool",
        "body": (
"1. Interface overview\n"
"· Top toolbar: Build Mod | Generate PNG from Font | Batch Replace | Number Size | Tutorial | Language toggle (中文/EN)\n"
"· Second row: Mod Name (editable) + Mod Output Dir (default ./mod_output) + Browse\n"
"· Table 5 columns: Name | Coord | Native (real ratio) | Action (Keep/Block/Replace) | Replace Image (preview + import + scale + stretch)\n"
"· Column header of Replace Image also has: Batch Import from Folder... + Scale + Stretch\n"
"· Under column headers ①②③ there is an 'Export…' button each (name list / coords / native PNGs); toolbar 'Export all cols…' exports the whole table to Excel\n"
"· Table grouped & collapsible (Base Add / Damage / Effect / Flesh Weak 1 / Flesh Weak 2 / Heal / Element Counter); click group title to expand/collapse\n"
"· Bottom: Console (log)\n"
"\n"
"2. Replace glyphs (3 ways - use any, or mix)\n"
"1) Generate colored PNGs from a font (TTF/OTF) (easiest)\n"
"   · Click 'Generate PNG from Font' -> dialog\n"
"   · Select font: in 'Font File' row click Browse and choose your .ttf or .otf (default ./LiXuKeShuFa-1.ttf; swap for any font)\n"
"   · Font PNG output dir: the dialog's 'Output Dir' (default ./font_output) is where generated PNGs are saved, relative to the exe folder (run exe directly -> font_output created next to it). After generation these PNGs are auto-imported as Replace\n"
"   · Per-group color / opacity / bold / italic: customize per group (default = each group's native average color; Element Counter defaults to gold)\n"
"   · Click 'Generate PNG' to start; a popup reports count and path when done\n"
"2) Manual single import: in Replace Image column click a row's Import.. and pick a PNG -> auto set to Replace; adjust Scale / Stretch\n"
"3) Batch import from folder: in Replace Image header click 'Batch Import from Folder...' and pick a folder; files whose name equals the element name (e.g. hud_nmbtl_damage00.png) are auto-matched and set as Replace; Scale/Stretch come from the header inputs\n"
"4) Batch replace by code (00-12): toolbar 'Batch Replace' dialog lists 13 codes; supply one image per code, click 'Apply to Group' to replace all sprites with that code across groups\n"
"\n"
"3. Number appearance / position / teammate color (optional)\n"
"· Click 'Damage Size / Float / Teammate Color' -> grouped input boxes:\n"
"   [Position & Time] height_: overall vertical pos of damage area, 0=horizontal axis through screen center recommend -2~2; length_: overall horizontal pos, 0=vertical axis through screen center recommend -200~200; viewTime_: display time(sec) recommend 0.2\n"
"   [Spacing] space_: character crowding per damage string recommend -40~40; healSpace_: character crowding per heal string recommend -40~40\n"
"   [Size] commonSize_(non-crit ally) recommend 0.2~1; criticalSize_(crit ally) recommend 0.2~1; playerSize_(non-crit you) recommend 0.3~1.2; playerCriticalSize_(crit you) recommend 0.3~1.2; spArtsLinkAttackSize_(elemental advantage scale) 0.3~2(0=off,1=native)\n"
        "   [Float] Spread factor alpha: one value scales all 8 native values of rangeNear_/rangeFar_; 1.0=native, larger=more spread, smaller=more compact; suggest 0.5~3.0\n"
"   [Teammate color RGBA] normalAttackColor_ (no counter) / spAttackColor_ (counter): click the swatch to open the palette, or fill R/G/B/A (0-255, A=opacity)\n"
"· Type a float; Enter or focus-out applies it. 'Reset All' restores the native base\n"
"\n"
"4. Export (optional)\n"
"· Under column headers ①②③ there is an 'Export…' button each; the save dialog is pre-filled with a language-based default name, so you can save directly or rename:\n"
"   - Col 1 (Name): exports Excel 'element_names.xlsx' (index + name)\n"
"   - Col 2 (Coord): exports Excel 'coordinates.xlsx' (index + name + X + Y + W + H)\n"
"   - Col 3 (Native): pick a folder, each native image saved as 'name.png' (opens 'native_pngs' folder by default)\n"
"· Toolbar 'Export all cols…' (right of Number Size): exports the whole table to Excel with thumbnails (name / coord / size / native image / action / replace image)\n"
"· Default names differ by language: English as listed above; Chinese are 元素名列表 / 坐标表 / 全列导出 / 原生图\n"
"\n"
"5. Build Mod\n"
"· Set Mod Output Dir (default ./mod_output - NOT the same as the font PNG output dir)\n"
"· Set Mod Name (optional)\n"
"· Click 'Build Mod' -> creates the zip and opens its folder\n"
"\n"
"6. Install\n"
"· Requires Reloaded-II + gbfrelink.utility.manager; load the generated mod and launch the game\n"
"\n"
"Note: don't confuse the two 'output dirs'\n"
"· Top Mod Output Dir = final mod zip (default ./mod_output)\n"
"· Inside 'Generate PNG from Font' dialog, Output Dir = font PNGs (default ./font_output, relative to exe dir)\n"
        ),
    },
}


def tr(key, *args):
    d = I18N.get(LANG, I18N["zh"])
    s = d.get(key)
    if s is None:
        s = I18N["zh"].get(key, key)
    if args:
        try:
            return s % args
        except Exception:
            return s
    return s


def now():
    return datetime.datetime.now().strftime("%H:%M:%S")


def app_base():
    """返回程序运行根目录: 打包后为 exe 所在目录, 开发期为本文件所在目录。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def app_root():
    """所有浏览对话框的默认目录的统一来源: 始终指向 exe 所在的"真正的根目录"。
    若 exe 运行在 mod_output 子目录内, 自动回到上一级(真正的根目录),
    无论用户双击哪个副本, 浏览框都从根目录打开。"""
    base = app_base()
    if os.path.basename(base).lower() == "mod_output":
        parent = os.path.dirname(base)
        if os.path.isdir(parent):
            return parent
    return base


def checkerboard(size, sq=8):
    w, h = size
    img = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    px = img.load()
    for y in range(h):
        for x in range(w):
            c = ((x // sq) + (y // sq)) % 2
            col = (90, 90, 90, 255) if c else (150, 150, 150, 255)
            px[x, y] = col
    return img


def to_thumb(pil, max_sz, checker=True):
    """pil: RGBA. 保持比例缩略; 透明处垫棋盘格。返回 PhotoImage。"""
    w, h = pil.size
    scale = min(max_sz[0] / w, max_sz[1] / h, 1.0)
    if scale < 1.0:
        pil = pil.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.LANCZOS)
    if checker:
        bg = checkerboard(pil.size)
        bg.paste(pil, (0, 0), pil)
        pil = bg
    return ImageTk.PhotoImage(pil)


core.set_translator(tr)
class App:
    def __init__(self, root):
        self.root = root
        self.root.title(tr("app_title"))
        self.root.geometry("1020x780")
        self._i18n = []   # 语言切换时刷新的控件注册表: (widget, key, attr)

        self.native = core.load_native()
        self.sprites = self.native["sprites"]
        self.hd_arr = self.native["hd_arr"]
        # 分组(按出现顺序)
        self.groups = self._compute_groups()
        # 每组默认颜色 = 该组可替换精灵原生主色均值(忠实于当前配色)
        self.group_colors = self._default_group_colors()
        # TTF 每组独立样式: {group: {bold, italic}}
        self.group_styles = {}
        # TTF 每组不透明度 (0-255, 默认255=完全不透明)
        self.group_opacities = {}
        for g in self.groups:
            self.group_styles[g] = {"bold": False, "italic": False}
            self.group_opacities[g] = 255
        # 伤害数字参数(hud_param.json): 默认 = 原生游戏基准值
        self.dmg_sizes = dict(DEFAULT_SIZE_SCALE)
        self.dmg_colors = {k: list(v) for k, v in DEFAULT_COLOR.items()}
        self.dmg_range_alpha = DEFAULT_RANGE_ALPHA
        # 每精灵运行态
        self.rows = []          # list of dict
        self.row_by_name = {}
        self.photo_refs = []    # 防止 PhotoImage 被 GC

        self._build_toolbar()
        self._build_header()
        self._build_table()
        self._build_console()
        self._bind_wheel()

        self.log(tr("log_loaded")
                 % (self.native["AW"], self.native["AH"], self.native["FW"], self.native["FH"],
                    len(self.sprites), sum(1 for s in self.sprites if s["char"])))
        self.log(tr("log_tip"))

    # ───────────────── 工具栏 ─────────────────
    def _build_toolbar(self):
        bar = ttk.Frame(self.root)
        bar.pack(side="top", fill="x")
        b1 = ttk.Button(bar, text=tr("gen_mod"), command=self.on_build)
        b1.pack(side="left", padx=2); self._reg(b1, "gen_mod")
        b2 = ttk.Button(bar, text=tr("ttf_tab"), command=self.on_ttf)
        b2.pack(side="left", padx=2); self._reg(b2, "ttf_tab")
        b3 = ttk.Button(bar, text=tr("batch_replace"), command=self.on_batch_replace)
        b3.pack(side="left", padx=2); self._reg(b3, "batch_replace")
        b4 = ttk.Button(bar, text=tr("size_control"), command=self.on_size_control)
        b4.pack(side="left", padx=2); self._reg(b4, "size_control")
        b4e = ttk.Button(bar, text=tr("exp_all"), command=self.on_export_all)
        b4e.pack(side="left", padx=2); self._reg(b4e, "exp_all")
        # 右上角: 中英切换
        self._lang_btn = ttk.Button(bar, text=tr("lang_toggle"), command=self._on_toggle_lang, width=5)
        self._lang_btn.pack(side="right", padx=6)
        # 教程按钮: 紧挨语言切换左边
        self._tut_btn = ttk.Button(bar, text=tr("tutorial"), command=self.on_tutorial)
        self._tut_btn.pack(side="right", padx=6); self._reg(self._tut_btn, "tutorial")

        sub = ttk.Frame(self.root)
        sub.pack(side="top", fill="x")
        l1 = ttk.Label(sub, text=tr("mod_name"))
        l1.pack(side="left", padx=2); self._reg(l1, "mod_name")
        self.mod_name = tk.StringVar(value="GBFR Damage Font Editor Mod")
        ttk.Entry(sub, textvariable=self.mod_name, width=30).pack(side="left", padx=2)
        l2 = ttk.Label(sub, text=tr("mod_out"))
        l2.pack(side="left", padx=4); self._reg(l2, "mod_out")
        self.out_dir = tk.StringVar(value="./mod_output")
        ttk.Entry(sub, textvariable=self.out_dir, width=34).pack(side="left", padx=2)
        b5 = ttk.Button(sub, text=tr("browse"), command=self.on_browse_out)
        b5.pack(side="left", padx=2); self._reg(b5, "browse")

    # ───────────────── 列头 ─────────────────
    def _build_header(self):
        hdr = ttk.Frame(self.root)
        hdr.pack(side="top", fill="x")
        titles = [tr("col_name"), tr("col_coord"), tr("col_native"), tr("col_action"), tr("col_replace")]
        cols = ["col_name", "col_coord", "col_native", "col_action", "col_replace"]
        export_keys = {COL_NAME: "exp_col", COL_COORD: "exp_col", COL_NATIVE: "exp_col"}
        # Use grid layout so columns align with the table below
        for c in range(5):
            hdr.grid_columnconfigure(c, minsize=COL_W[c], weight=(1 if c == COL_REPLACE else 0))
        h_row = 0
        # Row 0: column title labels
        for c in range(5):
            lab = ttk.Label(hdr, text=titles[c], font=("Microsoft YaHei", 12, "bold"), anchor="center")
            lab.grid(row=h_row, column=c, sticky="ew", padx=1, pady=1)
            self._reg(lab, cols[c])
        h_row += 1
        # Row 1: sub-controls under each column
        for c in range(5):
            if c == COL_ACTION:
                f = ttk.Frame(hdr)
                f.grid(row=h_row, column=c, sticky="ew", padx=1, pady=1)
                self._bulk_action_frame(f)
            elif c == COL_REPLACE:
                f = ttk.Frame(hdr)
                f.grid(row=h_row, column=c, sticky="ew", padx=1, pady=1)
                self._batch_import_frame(f)
            elif c in export_keys:
                eb = ttk.Button(hdr, text=tr("exp_col"), command=lambda cc=c: self.on_export_col(cc))
                eb.grid(row=h_row, column=c, sticky="ew", padx=1, pady=(0, 2)); self._reg(eb, "exp_col")

    def _bulk_action_frame(self, f):
        container = ttk.Frame(f)
        container.pack(fill="x", pady=1)
        t = ttk.Frame(container)
        t.pack(anchor="center")
        self.bulk_vars = {k: tk.BooleanVar() for k in ("keep", "block", "replace")}
        def mk(k, key):
            def cb():
                for kk in self.bulk_vars:
                    self.bulk_vars[kk].set(kk == k)
                self.bulk_action(k)
            cb2 = lambda kk=k: cb()
            cbw = ttk.Checkbutton(t, text=tr(key), variable=self.bulk_vars[k], command=cb2)
            cbw.pack(side="left", padx=3); self._reg(cbw, key)
        mk("keep", "all_keep"); mk("block", "all_block"); mk("replace", "all_replace")

    def _batch_import_frame(self, f):
        container = ttk.Frame(f)
        container.pack(fill="x", pady=1)
        t = ttk.Frame(container)
        t.pack(anchor="center")
        bi = ttk.Button(t, text=tr("batch_import"), command=self.on_batch_import)
        bi.pack(side="left", padx=2); self._reg(bi, "batch_import")
        rl = ttk.Label(t, text=tr("ratio"))
        rl.pack(side="left", padx=(4, 0)); self._reg(rl, "ratio")
        self.batch_ratio = tk.IntVar(value=100)
        ttk.Entry(t, textvariable=self.batch_ratio, width=6).pack(side="left", padx=1)
        ttk.Label(t, text="%").pack(side="left")
        self.batch_stretch = tk.BooleanVar(value=False)
        sc = ttk.Checkbutton(t, text=tr("stretch"), variable=self.batch_stretch)
        sc.pack(side="left", padx=(4, 0)); self._reg(sc, "stretch")

    # ───────────────── 分组辅助 ─────────────────
    def _compute_groups(self):
        """按精灵出现顺序去重得到组列表。"""
        order, seen = [], set()
        for sp in self.sprites:
            g = sp["group"]
            if g not in seen:
                seen.add(g)
                order.append(g)
        return order

    def _group_label(self, g):
        """返回组别的 UI 显示名。"""
        k = GROUP_TR.get(g)
        return tr(k) if k else g

    # ───────────────── 语言切换 ─────────────────
    def _reg(self, widget, key, attr="text"):
        """注册一个需要随语言切换刷新的控件。"""
        self._i18n.append((widget, key, attr))
        return widget

    def _apply_lang(self):
        """刷新所有已注册控件 + 窗口标题 + 组头。"""
        self.root.title(tr("app_title"))
        for widget, key, attr in self._i18n:
            try:
                widget[attr] = tr(key)
            except Exception:
                pass
        self._refresh_all_group_labels()
        if getattr(self, "_lang_btn", None) is not None:
            try:
                self._lang_btn["text"] = tr("lang_toggle")
            except Exception:
                pass

    def _refresh_all_group_labels(self):
        for g in self.group_states:
            self._refresh_group_label(g)

    def _on_toggle_lang(self):
        global LANG
        LANG = "en" if LANG == "zh" else "zh"
        self._apply_lang()

    def _default_group_colors(self):
        """每组默认颜色 = 该组可替换精灵原生主色(r,g,b)均值(忠实于当前配色)。"""
        res = {}
        for g in self.groups:
            cols = [tuple(int(c) for c in s["dominant"]) for s in self.sprites
                    if s["group"] == g and s.get("char")]
            if cols:
                arr = np.array(cols, dtype=float)
                res[g] = tuple(int(round(v)) for v in arr.mean(axis=0))
            else:
                res[g] = (255, 255, 255)
        # 覆盖 link 组默认颜色为金黄色
        res["link"] = (253, 201, 19)
        return res

    # ───────────────── 表格 ─────────────────
    def _build_table(self):
        wrap = ttk.Frame(self.root)
        wrap.pack(side="top", fill="both", expand=True)
        self.canvas = tk.Canvas(wrap)
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.table = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.table, anchor="nw")
        self.canvas.bind("<Button-1>", self._on_canvas_click)
        self.table.bind("<Configure>", lambda e: self._update_scrollregion())
        for c in range(5):
            self.table.grid_columnconfigure(c, minsize=COL_W[c])
        self.table.grid_columnconfigure(4, weight=1)  # 末列拉伸填满宽度
        self.group_states = {}
        self._next_row = 0
        for g in self.groups:
            sps = [s for s in self.sprites if s["group"] == g]
            self._build_group_header(g, sps)
            for sp in sps:
                self._build_row(sp)

    def _build_group_header(self, g, sps):
        row = self._next_row
        self._next_row += 1
        n_char = sum(1 for s in sps if s.get("char"))
        f = tk.Frame(self.table, bg="#37506b", relief="raised", bd=1)
        f.grid(row=row, column=0, columnspan=5, sticky="ew", padx=1, pady=2)
        lbl = tk.Label(f, text="", bg="#37506b", fg="white",
                       font=("Microsoft YaHei", 10, "bold"), anchor="w",
                       padx=8, pady=4, cursor="hand2")
        lbl.pack(side="left", fill="x", expand=True)
        f.bind("<Button-1>", lambda e, g=g: self._toggle_group(g))
        lbl.bind("<Button-1>", lambda e, g=g: self._toggle_group(g))
        self.group_states[g] = {"expanded": True, "header": f, "label": lbl,
                                "n_total": len(sps), "n_char": n_char, "rows": []}
        self._refresh_group_label(g)

    def _toggle_group(self, g):
        stt = self.group_states[g]
        stt["expanded"] = not stt["expanded"]
        if stt["expanded"]:
            for st in stt["rows"]:
                for cell in st["cells"]:
                    cell.grid()
        else:
            for st in stt["rows"]:
                for cell in st["cells"]:
                    cell.grid_remove()
        self._refresh_group_label(g)
        self._update_scrollregion()

    def _on_canvas_click(self, event):
        # 画布级兜底: 按点击 y 坐标定位组标题并收起/展开
        # (防止个别环境下嵌入控件命中测试失效导致点标题栏无反应)
        try:
            cy = self.canvas.canvasy(event.y)
        except Exception:
            return
        for g, stt in self.group_states.items():
            h = stt.get("header")
            if not h:
                continue
            y0 = h.winfo_y()
            y1 = y0 + h.winfo_height()
            if y0 <= cy <= y1:
                self._toggle_group(g)
                return

    def _update_scrollregion(self):
        """更新滚动区域 (V15 简单版本)。"""
        self.canvas.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _refresh_group_label(self, g):
        stt = self.group_states[g]
        ind = "▾" if stt["expanded"] else "▸"
        base = "%s  ·  %s" % (self._group_label(g), tr("grp_elems") % stt["n_total"])
        if stt["n_char"]:
            base += tr("grp_replaceable") % stt["n_char"]
        stt["label"].configure(text=ind + "  " + base)

    def _build_row(self, sp):
        row = self._next_row
        self._next_row += 1
        # 运行态
        st = {
            "sp": sp, "name": sp["name"],
            "action": "keep", "ratio": 100.0, "repl": None,
            "stretch": False,
            "native_photo": None, "preview_photo": None,
            "preview_label": None, "action_states": None,
            "cells": [], "grid_row": row,
        }
        # 原生裁图 (HD)
        crop = self.hd_arr[sp["y"]:sp["y"] + sp["h"], sp["x"]:sp["x"] + sp["w"]]
        st["native_pil"] = Image.fromarray(crop).convert("RGBA")

        for c in range(5):
            cell = ttk.Frame(self.table, width=COL_W[c], height=104)
            cell.grid(row=row, column=c, sticky="nsew", padx=1, pady=1)
            cell.grid_propagate(False)
            if c == COL_NAME:
                ttk.Label(cell, text=sp["name"], font=("Consolas", 8),
                          wraplength=COL_W[c] - 6, justify="center", anchor="center").pack(expand=True)
            elif c == COL_COORD:
                ttk.Label(cell, text="(%d,%d)\n%dx%d" % (sp["x"], sp["y"], sp["w"], sp["h"]),
                          font=("Consolas", 8), anchor="center").pack(expand=True)
            elif c == COL_NATIVE:
                ph = to_thumb(st["native_pil"], NATIVE_MAX)
                st["native_photo"] = ph
                self.photo_refs.append(ph)
                lab = ttk.Label(cell, image=ph)
                lab.pack(expand=True)
            elif c == COL_ACTION:
                self._action_cell(cell, st)
            elif c == COL_REPLACE:
                self._replace_cell(cell, st)
            st["cells"].append(cell)
        self.rows.append(st)
        self.row_by_name[sp["name"]] = st
        self.group_states[sp["group"]]["rows"].append(st)

    def _action_cell(self, cell, st):
        states = {k: tk.BooleanVar(value=(k == "keep")) for k in ("keep", "block", "replace")}
        st["action_states"] = states

        def sync(picked):
            for k in states:
                states[k].set(k == picked)
            st["action"] = picked
            self._refresh_preview(st)

        f = ttk.Frame(cell)
        f.pack(expand=True)
        for key, k in [("act_keep", "keep"), ("act_block", "block"), ("act_replace", "replace")]:
            cbw = ttk.Checkbutton(f, text=tr(key), variable=states[k],
                                  command=lambda kk=k: sync(kk))
            cbw.pack(side="left", padx=2); self._reg(cbw, key)

    def _replace_cell(self, cell, st):
        # 水平布局: 左侧预览图 | 右侧导入+占比控件
        outer = ttk.Frame(cell)
        outer.pack(expand=True, fill="both", padx=2, pady=2)

        # 左: 预览图 (用 place 强制居中, 兼容所有图片尺寸)
        left = ttk.Frame(outer)
        left.pack(side="left", expand=True, fill="both")
        prev = ttk.Label(left, text=tr("after_replace"), foreground="#888",
                         font=("Microsoft YaHei", 8), anchor="center")
        st["preview_label"] = prev
        self._reg(prev, "after_replace")
        # place 保证无论图片多大都居中显示
        prev.place(relx=0.5, rely=0.5, anchor="center")

        # 右: 导入 + 占比
        right = ttk.Frame(outer)
        right.pack(side="right", padx=(4, 0))

        btn = ttk.Button(right, text=tr("import"), width=6,
                         command=lambda s=st: self.on_import_one(s))
        btn.pack(pady=(0, 2)); self._reg(btn, "import")

        rf = ttk.Frame(right)
        rf.pack()
        rl = ttk.Label(rf, text=tr("ratio_lbl"))
        rl.pack(side="left", padx=(0, 2)); self._reg(rl, "ratio_lbl")
        rv = tk.IntVar(value=100)
        st["ratio_var"] = rv
        sv = tk.BooleanVar(value=False)
        st["stretch_var"] = sv

        def on_ratio(*a):
            try:
                v = int(rv.get())
            except Exception:
                v = 100
            v = max(0, v)   # 无上限, 输入多大用多大; 下限 0
            st["ratio"] = float(v)
            self._refresh_preview(st)

        def on_stretch(*a):
            st["stretch"] = sv.get()
            self._refresh_preview(st)

        rv.trace_add("write", on_ratio)
        sv.trace_add("write", on_stretch)
        ttk.Entry(rf, textvariable=rv, width=5).pack(side="left", padx=1)
        ttk.Label(rf, text="%").pack(side="left")
        sc = ttk.Checkbutton(rf, text=tr("stretch_lbl"), variable=sv)
        sc.pack(side="left", padx=(4, 0)); self._reg(sc, "stretch_lbl")

    def _refresh_preview(self, st):
        if st["action"] == "replace" and st["repl"] is not None:
            # 不传native: 预览展示游戏内实际效果(透明底+替换图, 原生已清除)
            box = core.composite_preview(st["sp"], st["repl"], st["ratio"], fhd=False,
                                         stretch=st.get("stretch", False))
            ph = to_thumb(box, PREVIEW_MAX)
            st["preview_photo"] = ph
            st["preview_label"].configure(image=ph, text="")
            st["preview_label"].image = ph
        else:
            st["preview_label"].configure(image="", text=tr("after_replace"))
            st["preview_label"].image = None

    # ───────────────── 控制台 ─────────────────
    def _build_console(self):
        f = ttk.Frame(self.root)
        f.pack(side="bottom", fill="x")
        cl = ttk.Label(f, text=tr("console"))
        cl.pack(anchor="w", padx=4); self._reg(cl, "console")
        self.console = tk.Text(f, height=8, bg="#1e1e1e", fg="#d0d0d0",
                               font=("Consolas", 9), state="disabled")
        self.console.pack(side="left", fill="both", expand=True, padx=4, pady=2)
        sb = ttk.Scrollbar(f, command=self.console.yview)
        sb.pack(side="right", fill="y")
        self.console.configure(yscrollcommand=sb.set)

    def log(self, msg):
        self.console.configure(state="normal")
        self.console.insert("end", "[%s] %s\n" % (now(), msg))
        self.console.see("end")
        self.console.configure(state="disabled")

    def _bind_wheel(self):
        self._bind_wheel_to(self.canvas)

    def _bind_wheel_to(self, canvas):
        """为某个可滚动 canvas 绑定滚轮。

        关键修复: 把滚轮绑定挂到 canvas 所属的 Toplevel 窗口上（主界面挂 root，
        弹窗挂各自 Toplevel），而不是 bind_all（全局 all 标签）。

        原因: bind_all 会把回调塞进全局 all 标签，弹窗关闭时必须手动 unbind_all
        清理；但本机 Python 3.13 的 tkinter 中 unbind_all 只接受 sequence 一个参数，
        多传 funcid 会抛 TypeError，导致弹窗滚轮绑定泄漏累积、并干扰主界面滚轮。
        改为挂到 Toplevel 后，该窗口内所有子控件共享此 bindtag，窗口销毁时绑定
        自动随之移除，零手动清理、零泄漏、零串扰。

        判定: 仅当滚轮事件落在 canvas 子树内（沿父链回溯命中 canvas）才滚动它，
        因此弹窗滚轮不会误滚主界面，主界面滚轮也不会误滚弹窗。"""
        win = canvas.winfo_toplevel()
        root = self.root

        def on_mousewheel(event):
            w = event.widget
            while w is not None:
                if w is canvas:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    return
                # 到达所属窗口仍没命中 canvas → 事件落在该窗口其它区域, 不滚动
                if w is win:
                    return
                p = w.winfo_parent()
                w = root.nametowidget(p) if p else None

        win.bind("<MouseWheel>", on_mousewheel)

    # ───────────────── 动作 ─────────────────
    def bulk_action(self, action):
        for st in self.rows:
            st["action"] = action
            if st["action_states"]:
                for k in st["action_states"]:
                    st["action_states"][k].set(k == action)
            self._refresh_preview(st)
        self.log(tr("log_bulk") % action)

    def on_browse_out(self):
        d = filedialog.askdirectory(title=tr("browse_out_title"), initialdir=app_root())
        if d:
            self.out_dir.set(d)

    def on_import_one(self, st):
        p = filedialog.askopenfilename(
            title=tr("br_import"), filetypes=[("PNG", "*.png"), ("Image", "*.png;*.jpg;*.bmp")],
            initialdir=app_root())
        if not p:
            return
        try:
            img = Image.open(p).convert("RGBA")
        except Exception as e:
            messagebox.showerror(tr("import_fail_title"), str(e))
            return
        st["repl"] = img
        # 自动切到替换
        st["action"] = "replace"
        for k in st["action_states"]:
            st["action_states"][k].set(k == "replace")
        self._refresh_preview(st)
        self.log(tr("log_import") % (st["name"], os.path.basename(p)))

    def on_batch_import(self):
        # 默认目录 = exe 根目录(统一来源 app_root, 含 mod_output 回退)
        d = filedialog.askdirectory(title=tr("browse_batch_title"), initialdir=app_root())
        if not d:
            return
        try:
            ratio = max(0, int(self.batch_ratio.get()))   # 无上限, 输入多大用多大
        except Exception:
            ratio = 100
        # 大小写不敏感的运行态索引(只匹配运行态 st dict, 而非元数据 dict)
        row_by_lower = {name.lower(): st for name, st in self.row_by_name.items()}
        n = 0
        skipped_effect = 0
        for fn in os.listdir(d):
            if not fn.lower().endswith(".png"):
                continue
            base = os.path.splitext(fn)[0]
            st = row_by_lower.get(base.lower())
            if not st:
                continue
            if st["sp"]["group"] == "effect":
                skipped_effect += 1
                continue
            try:
                img = Image.open(os.path.join(d, fn)).convert("RGBA")
            except Exception:
                continue
            st["repl"] = img
            st["action"] = "replace"
            st["ratio"] = ratio
            st["ratio_var"].set(ratio)
            stretch_val = self.batch_stretch.get()
            st["stretch"] = stretch_val
            if "stretch_var" in st:
                st["stretch_var"].set(stretch_val)
            if st["action_states"]:
                for k in st["action_states"]:
                    st["action_states"][k].set(k == "replace")
            self._refresh_preview(st)
            n += 1
        self.log(tr("log_batch") % (n, ratio))
        if skipped_effect:
            self.log(tr("log_batch_skip_effect") % skipped_effect)
        if n == 0:
            messagebox.showinfo(tr("batch_no_png_title"), tr("batch_no_png_msg"))

    def on_ttf(self):
        self._open_ttf_settings()

    # 固定 13 类编号: 符号 -> 2位编号
    CATEGORIES = [
        ("0", "00"), ("1", "01"), ("2", "02"), ("3", "03"), ("4", "04"),
        ("5", "05"), ("6", "06"), ("7", "07"), ("8", "08"), ("9", "09"),
        ("！", "10"), ("%", "11"), ("+", "12"),
    ]

    def _sprite_code(self, name):
        """从精灵名提取编号: 取最后一个2位数(如 hud_nmbtl_guard01_00 -> 00)。"""
        m = re.findall(r"\d{2}", name)
        return m[-1] if m else None

    def on_batch_replace(self):
        """批量替换图片: 13 类编号(00-12)同时展示, 每类导入一张图+占比+铺满, 点'组内同类别覆盖'替换该编号在各组的所有精灵。"""
        win = tk.Toplevel(self.root)
        win.title(tr("br_win_title"))
        win.geometry("1089x922")
        win.transient(self.root)
        win.grab_set()

        ttk.Label(win, text=tr("br_desc"),
                  font=("Microsoft YaHei", 9), wraplength=1000, anchor="w").pack(pady=6, padx=10, fill="x")

        # 可滚动容器
        canvas = tk.Canvas(win)
        vsb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._bind_wheel_to(canvas)

        inner = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        for sym, code in self.CATEGORIES:
            # 该编号出现在哪些组
            groups_with = sorted(set(s["group"] for s in self.sprites
                                     if self._sprite_code(s["name"]) == code))
            glabels = "、".join(self._group_label(g) for g in groups_with) or tr("br_none")
            rowf = ttk.LabelFrame(inner, text=tr("br_row_frame") % (sym, code, glabels))
            rowf.pack(fill="x", padx=10, pady=4)

            # 左: 预览
            lf = ttk.Frame(rowf)
            lf.pack(side="left", padx=4)
            prev = ttk.Label(lf, text=tr("br_unimported"), foreground="#888")
            prev.pack()

            # 右: 控件
            rf = ttk.Frame(rowf)
            rf.pack(side="left", padx=8, fill="x", expand=True)

            ratio_var = tk.IntVar(value=100)
            stretch_var = tk.BooleanVar(value=False)
            img_ref = [None]
            preview_ref = [None]

            def make_pick(prev_lbl):
                def _pick(prev_lbl=prev_lbl, img_ref=img_ref):
                    p = filedialog.askopenfilename(
                        title=tr("br_import"),
                        filetypes=[("PNG", "*.png"), ("Image", "*.png;*.jpg;*.bmp")],
                        initialdir=app_root())
                    if not p:
                        return
                    try:
                        img = Image.open(p).convert("RGBA")
                    except Exception as e:
                        messagebox.showerror(tr("import_fail_title"), str(e))
                        return
                    img_ref[0] = img
                    thumb = to_thumb(img, (110, 70))
                    preview_ref[0] = thumb
                    prev_lbl.configure(image=thumb, text="")
                    prev_lbl.image = thumb
                return _pick

            def make_apply():
                def _apply(code=code, img_ref=img_ref, ratio_var=ratio_var, stretch_var=stretch_var):
                    if img_ref[0] is None:
                        messagebox.showwarning(tr("br_warn"), tr("br_warn_msg"))
                        return
                    try:
                        ratio_val = max(0, int(ratio_var.get()))
                    except Exception:
                        ratio_val = 100
                    stretch_val = stretch_var.get()
                    applied = 0
                    for st in self.rows:
                        if st["sp"]["group"] == "effect":
                            continue
                        if self._sprite_code(st["name"]) == code:
                            st["repl"] = img_ref[0].copy()
                            st["action"] = "replace"
                            st["ratio"] = float(ratio_val)
                            st["ratio_var"].set(ratio_val)
                            st["stretch"] = stretch_val
                            if "stretch_var" in st:
                                st["stretch_var"].set(stretch_val)
                            for k in st["action_states"]:
                                st["action_states"][k].set(k == "replace")
                            self._refresh_preview(st)
                            applied += 1
                    mode_str = tr("stretch_lbl") if stretch_val else (tr("br_scale_mode") % ratio_val)
                    self.log(tr("br_cover_log") % (code, mode_str, applied))
                    messagebox.showinfo(tr("br_done"),
                                        tr("br_done_msg") % (code, applied))
                return _apply

            ttk.Button(rf, text=tr("br_import_btn"), command=make_pick(prev)).pack(side="left", padx=2)

            rrow = ttk.Frame(rf)
            rrow.pack(side="left", padx=4)
            ttk.Label(rrow, text=tr("ratio_lbl")).pack(side="left")
            ttk.Entry(rrow, textvariable=ratio_var, width=5).pack(side="left")
            ttk.Label(rrow, text="%").pack(side="left")
            ttk.Checkbutton(rrow, text=tr("stretch_lbl"), variable=stretch_var).pack(side="left", padx=(4, 0))

            ttk.Button(rf, text=tr("br_apply"), command=make_apply()).pack(side="left", padx=6)

    def on_size_control(self):
        """Number size/pos/spacing + teammate color control (hud_param.json damage_)."""
        win = tk.Toplevel(self.root)
        win.title(tr("sz_win_title"))
        win.geometry("720x780")
        win.resizable(True, True)
        win.transient(self.root)
        win.grab_set()

        # scrollable canvas for mouse wheel
        cvas = tk.Canvas(win, highlightthickness=0)
        vsb = ttk.Scrollbar(win, orient="vertical", command=cvas.yview)
        scroll_frame = ttk.Frame(cvas)
        scroll_frame.bind("<Configure>", lambda e: cvas.configure(scrollregion=cvas.bbox("all")))
        cvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        cvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        cvas.pack(side="left", fill="both", expand=True)
        def _on_mousewheel(event):
            cvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        win.bind("<MouseWheel>", _on_mousewheel)
        _p = scroll_frame

        # top description
        desc_frm = ttk.Frame(_p)
        desc_frm.pack(fill="x", padx=16, pady=(10, 4))
        ttk.Label(desc_frm, text=tr("sz_desc"),
                  font=("Microsoft YaHei", 9), wraplength=520, justify="left").pack(anchor="w")

        sep = ttk.Separator(_p, orient="horizontal")
        sep.pack(fill="x", padx=16, pady=6)

        # ── 数值分组: 位置与时间 / 间距 / 大小 ──
        self._size_entries = {}
        groups = [
            (tr("sz_sec_pos"), ["height_", "length_", "viewTime_"]),
            (tr("sz_sec_space"), ["space_", "healSpace_"]),
            (tr("sz_sec_size"), ["commonSize_", "criticalSize_", "playerSize_", "playerCriticalSize_", "spArtsLinkAttackSize_"]),
        ]
        for sec_title, keys in groups:
            lf = ttk.LabelFrame(_p, text=sec_title)
            lf.pack(fill="x", padx=16, pady=4)
            grid = ttk.Frame(lf)
            grid.pack(fill="x", padx=8, pady=4)
            grid.grid_columnconfigure(1, weight=1)
            for i, k in enumerate(keys):
                lbl = ttk.Label(grid, text=tr(SIZE_LABEL_KEYS[k]),
                                font=("Microsoft YaHei", 10, "bold"))
                lbl.grid(row=i, column=0, sticky="e", padx=(0, 10), pady=6)
                ent = ttk.Entry(grid, width=9, justify="center", font=("Consolas", 11))
                ent.insert(0, "%.2f" % self.dmg_sizes.get(k, DEFAULT_SIZE_SCALE.get(k, 0.0)))
                ent.grid(row=i, column=1, sticky="w", padx=6, pady=6)
                ent.bind("<Return>", lambda e, kk=k: self._on_size_entry(kk))
                ent.bind("<FocusOut>", lambda e, kk=k: self._on_size_entry(kk))
                self._size_entries[k] = ent
                hint = ttk.Label(grid, text=tr(SIZE_HINT_KEYS[k]),
                                 font=("Microsoft YaHei", 8), foreground="#888")
                hint.grid(row=i, column=2, sticky="w", padx=(8, 0), pady=6)

        # 浮动系数 —— 单个 alpha 乘以 rangeNear_/rangeFar_ 各原生值
        lfr = ttk.LabelFrame(_p, text=tr("sz_sec_float"))
        lfr.pack(fill="x", padx=16, pady=4)
        rowf = ttk.Frame(lfr)
        rowf.pack(fill="x", padx=8, pady=6)
        ttk.Label(rowf, text=tr("sz_range_alpha"),
                  font=("Microsoft YaHei", 10, "bold")).pack(side="left", padx=(0, 8))
        self._range_alpha_entry = ttk.Entry(rowf, width=10, justify="center", font=("Consolas", 12))
        self._range_alpha_entry.insert(0, "%.2f" % self.dmg_range_alpha)
        self._range_alpha_entry.pack(side="left", padx=6)
        self._range_alpha_entry.bind("<Return>", lambda e: self._on_range_alpha_entry())
        self._range_alpha_entry.bind("<FocusOut>", lambda e: self._on_range_alpha_entry())
        ttk.Label(rowf, text=tr("sz_range_alpha_hint"),
                  font=("Microsoft YaHei", 8), foreground="#888",
                  wraplength=380).pack(side="left", padx=(8, 0))

        # 队友伤害颜色 (RGBA)
        lfc = ttk.LabelFrame(_p, text=tr("sz_sec_color"))
        lfc.pack(fill="x", padx=16, pady=4)
        self._dmg_color_widgets = {}
        for i, k in enumerate(COLOR_KEYS):
            col = tuple(self.dmg_colors.get(k, DEFAULT_COLOR[k]))
            rowf = ttk.Frame(lfc)
            rowf.pack(fill="x", padx=8, pady=4)
            ttk.Label(rowf, text=tr(COLOR_LABEL_KEYS[k]),
                      font=("Microsoft YaHei", 10, "bold")).pack(side="left", padx=(0, 8))
            sw = tk.Button(rowf, width=3, bg="#%02x%02x%02x" % col[:3], relief="raised")
            rgb_frame = ttk.Frame(rowf)
            rgb_vars = []
            rgb_entries = []
            for idx2, ch in enumerate(["R", "G", "B", "A"]):
                ttk.Label(rgb_frame, text=ch, font=("Consolas", 8, "bold"),
                          foreground="#666").pack(side="left")
                v = tk.StringVar(value=str(col[idx2]))
                e = ttk.Entry(rgb_frame, textvariable=v, width=4)
                e.pack(side="left", padx=1)
                rgb_vars.append(v)
                rgb_entries.append(e)
            sw.configure(command=self._make_dmg_color_setter(k, sw, rgb_entries))
            rgb_apply = self._make_dmg_color_rgba(k, sw, rgb_entries)
            for e in rgb_entries:
                e.bind("<Return>", lambda ev, fn=rgb_apply: fn())
                e.bind("<FocusOut>", lambda ev, fn=rgb_apply: fn())
            sw.pack(side="left", padx=3)
            rgb_frame.pack(side="left", padx=(0, 6))
            ttk.Label(rowf, text=tr(COLOR_HINT_KEYS[k]),
                      font=("Microsoft YaHei", 8), foreground="#888").pack(side="left", padx=(6, 0))
            self._dmg_color_widgets[k] = {"btn": sw, "rgb_vars": rgb_vars}
        bottom = ttk.Frame(_p)
        bottom.pack(fill="x", padx=16, pady=(8, 10))
        ttk.Label(bottom, text=tr("sz_note"),
                  font=("Microsoft YaHei", 8), foreground="#888").pack(side="left", anchor="w")
        btn_area = ttk.Frame(bottom)
        btn_area.pack(side="right")
        ttk.Button(btn_area, text=tr("sz_reset"),
                   command=self._on_size_reset_all).pack(side="left", padx=4)
        ttk.Button(btn_area, text=tr("sz_close"),
                   command=win.destroy).pack(side="left", padx=4)

    def _make_dmg_color_setter(self, k, btn, rgb_entries):
        def cb():
            cur = tuple(self.dmg_colors.get(k, DEFAULT_COLOR[k]))
            c = colorchooser.askcolor(color="#%02x%02x%02x" % cur[:3],
                                      title=tr("col_pick") % k)
            if c[1] is None:
                return
            r, g_, b = (int(x) for x in c[0])
            a = cur[3]
            self.dmg_colors[k] = (r, g_, b, a)
            btn.configure(bg="#%02x%02x%02x" % (r, g_, b))
            for idx2, val in enumerate([r, g_, b, a]):
                rgb_entries[idx2].delete(0, "end")
                rgb_entries[idx2].insert(0, str(val))
        return cb

    def _make_dmg_color_rgba(self, k, btn, rgb_entries):
        def apply():
            try:
                vals = [max(0, min(255, int(float(rgb_entries[i].get())))) for i in range(4)]
            except (ValueError, TypeError):
                return
            self.dmg_colors[k] = tuple(vals)
            btn.configure(bg="#%02x%02x%02x" % (vals[0], vals[1], vals[2]))
        return apply

    def _on_size_entry(self, k):
        ent = self._size_entries.get(k)
        if not ent:
            return
        raw = ent.get().strip()
        try:
            fv = float(raw)
        except Exception:
            fv = None
        if fv is None:
            fv = self.dmg_sizes.get(k, DEFAULT_SIZE_SCALE.get(k, 0.0))
        else:
            fv = round(fv, 2)
            if k == "spArtsLinkAttackSize_":
                if fv < 0:
                    fv = DEFAULT_SIZE_SCALE.get(k, 1.0)
        self.dmg_sizes[k] = fv
        ent.delete(0, "end")
        ent.insert(0, "%.2f" % fv)

    def _on_range_alpha_entry(self):
        raw = self._range_alpha_entry.get().strip()
        try:
            fv = float(raw)
        except Exception:
            fv = None
        if fv is None:
            fv = self.dmg_range_alpha
        else:
            fv = round(fv, 2)
        self.dmg_range_alpha = fv
        self._range_alpha_entry.delete(0, 'end')
        self._range_alpha_entry.insert(0, "%.2f" % fv)

    def _on_size_reset_all(self):
        for k in core.SIZE_KEYS:
            native_val = DEFAULT_SIZE_SCALE.get(k, 0.0)
            self.dmg_sizes[k] = native_val
            ent = self._size_entries.get(k)
            if ent:
                ent.delete(0, "end")
                ent.insert(0, "%.2f" % native_val)
        for k in COLOR_KEYS:
            col = DEFAULT_COLOR.get(k, (255, 255, 255, 255))
            self.dmg_colors[k] = tuple(col)
            gw = self._dmg_color_widgets.get(k)
            if gw:
                gw["btn"].configure(bg="#%02x%02x%02x" % col[:3])
                for idx2, val in enumerate(col):
                    gw["rgb_vars"][idx2].set(str(val))

        # reset range alpha to native (1.0)
        self.dmg_range_alpha = DEFAULT_RANGE_ALPHA
        if hasattr(self, '_range_alpha_entry'):
            self._range_alpha_entry.delete(0, 'end')
            self._range_alpha_entry.insert(0, "%.2f" % DEFAULT_RANGE_ALPHA)

    def _open_ttf_settings(self):
        """TTF 生成设置: 文件路径 + 每组颜色/加粗/斜体。"""
        win = tk.Toplevel(self.root)
        win.title(tr("ttf_win_title"))
        win.geometry("620x680")
        win.transient(self.root)
        win.grab_set()

        ttk.Label(win, text=tr("ttf_heading"), font=("Microsoft YaHei", 12, "bold")).pack(pady=6)

        # ── 文件选择区 ──
        ff = ttk.LabelFrame(win, text=tr("ttf_file"))
        ff.pack(fill="x", padx=10, pady=4)
        f1 = ttk.Frame(ff); f1.pack(fill="x", padx=4, pady=2)
        ttk.Label(f1, text=tr("ttf_font")).pack(side="left")
        self._dlg_ttf_path = tk.StringVar(value="./LiXuKeShuFa-1.ttf")
        ttk.Entry(f1, textvariable=self._dlg_ttf_path, width=40).pack(side="left", padx=4, fill="x", expand=True)
        ttk.Button(f1, text=tr("browse"), command=lambda: self._dlg_pick_ttf()).pack(side="left")

        f2 = ttk.Frame(ff); f2.pack(fill="x", padx=4, pady=2)
        ttk.Label(f2, text=tr("ttf_out")).pack(side="left")
        self._dlg_out_dir = tk.StringVar(value="./font_output")
        ttk.Entry(f2, textvariable=self._dlg_out_dir, width=40).pack(side="left", padx=4, fill="x", expand=True)
        ttk.Button(f2, text=tr("browse"), command=lambda: self._dlg_pick_outdir()).pack(side="left")

        # ── 各组颜色+样式 ── (可滚动)
        cf = ttk.LabelFrame(win, text=tr("ttf_groups"))
        cf.pack(fill="both", expand=True, padx=10, pady=4)

        canvas = tk.Canvas(cf, height=200)
        vsb = ttk.Scrollbar(cf, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._bind_wheel_to(canvas)

        inner = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # 存每组控件的引用(用于读取值 + 重置颜色)
        self._dlg_group_widgets = {}

        for g in self.groups:
            rowf = ttk.Frame(inner)
            rowf.pack(fill="x", padx=4, pady=3)

            # 组名(用显示名, 动态宽度, 格式: "组名 共N条")
            n = sum(1 for s in self.sprites if s["group"] == g)
            nc = sum(1 for s in self.sprites if s["group"] == g and s.get("char"))
            ttk.Label(rowf, text=self._group_label(g) + (tr("ttf_items") % n),
                      font=("Microsoft YaHei", 9, "bold")).pack(side="left")

            # 颜色色块
            col = self.group_colors[g]
            sw = tk.Button(rowf, width=3, bg="#%02x%02x%02x" % col, relief="raised")

            gw = {"color_btn": sw}
            self._dlg_group_widgets[g] = gw

            def make_color_setter(gg, btn, r_ent, g_ent, b_ent):
                def cb():
                    c = colorchooser.askcolor(color="#%02x%02x%02x" % self.group_colors[gg],
                                              title=tr("ttf_pick_color") % gg)
                    if c[1] is None:
                        return
                    r, g_, b = (int(x) for x in c[0])
                    self.group_colors[gg] = (r, g_, b)
                    btn.configure(bg="#%02x%02x%02x" % (r, g_, b))
                    r_ent.delete(0, "end"); r_ent.insert(0, str(r))
                    g_ent.delete(0, "end"); g_ent.insert(0, str(g_))
                    b_ent.delete(0, "end"); b_ent.insert(0, str(b))
                return cb

            def make_rgb_updater(gg, btn, r_ent, g_ent, b_ent):
                # read RGB entries and update color swatch
                def apply():
                    try:
                        r = max(0, min(255, int(float(r_ent.get()))))
                        g_ = max(0, min(255, int(float(g_ent.get()))))
                        b = max(0, min(255, int(float(b_ent.get()))))
                    except (ValueError, TypeError):
                        return
                    self.group_colors[gg] = (r, g_, b)
                    btn.configure(bg="#%02x%02x%02x" % (r, g_, b))
                return apply

            # RGB 可编辑输入框
            rgb_frame = ttk.Frame(rowf); rgb_frame.pack(side="left", padx=2)
            gw["rgb_vars"] = []
            rgb_entries = []
            for idx, ch in enumerate(["R", "G", "B"]):
                ttk.Label(rgb_frame, text=ch, font=("Consolas", 8, "bold"),
                         foreground="#666").pack(side="left")
                v = tk.StringVar(value=str(col[idx]))
                gw["rgb_vars"].append(v)
                e = ttk.Entry(rgb_frame, textvariable=v, width=4)
                e.pack(side="left", padx=1)
                rgb_entries.append(e)

            sw.configure(command=make_color_setter(g, sw, rgb_entries[0], rgb_entries[1], rgb_entries[2]))
            rgb_apply = make_rgb_updater(g, sw, rgb_entries[0], rgb_entries[1], rgb_entries[2])
            for e in rgb_entries:
                e.bind("<Return>", lambda ev, fn=rgb_apply: fn())
                e.bind("<FocusOut>", lambda ev, fn=rgb_apply: fn())
            sw.pack(side="left", padx=3)
            rgb_frame.pack(side="left", padx=(0, 6))

            # 不透明度
            opa_frame = ttk.Frame(rowf); opa_frame.pack(side="left", padx=2)
            ttk.Label(opa_frame, text=tr("ttf_opacity"), font=("Microsoft YaHei", 8)).pack(side="left")
            opa_var = tk.IntVar(value=int(self.group_opacities[g]))
            gw["opacity_var"] = opa_var
            opa_entry = ttk.Entry(opa_frame, textvariable=opa_var, width=4)
            opa_entry.pack(side="left", padx=1)
            ttk.Label(opa_frame, text="%", font=("Microsoft YaHei", 8)).pack(side="left")

            # 样式: 加粗 / 斜体
            gs = self.group_styles[g]
            gw["bold_var"] = tk.BooleanVar(value=gs["bold"])
            gw["italic_var"] = tk.BooleanVar(value=gs["italic"])

            ttk.Checkbutton(rowf, text=tr("ttf_bold"), variable=gw["bold_var"], width=5).pack(side="left", padx=2)
            ttk.Checkbutton(rowf, text=tr("ttf_italic"), variable=gw["italic_var"], width=5).pack(side="left", padx=2)

        # ── 底部按钮 ──
        btn_f = ttk.Frame(win)
        btn_f.pack(pady=8)

        def on_reset():
            self.group_colors = self._default_group_colors()
            for g in self.groups:
                self.group_styles[g] = {"bold": False, "italic": False}
                self.group_opacities[g] = 255
            # 重建组控件
            for child in inner.winfo_children():
                child.destroy()
            self._dlg_group_widgets.clear()
            for g in self.groups:
                rowf = ttk.Frame(inner)
                rowf.pack(fill="x", padx=4, pady=3)
                n = sum(1 for s in self.sprites if s["group"] == g)
                nc = sum(1 for s in self.sprites if s["group"] == g and s.get("char"))
                ttk.Label(rowf, text=self._group_label(g) + (tr("ttf_items") % n),
                          font=("Microsoft YaHei", 9, "bold")).pack(side="left")
                col = self.group_colors[g]
                sw = tk.Button(rowf, width=3, bg="#%02x%02x%02x" % col, relief="raised")
                gw = {"color_btn": sw}
                self._dlg_group_widgets[g] = gw

                def _make_cs(gg, btn, r_e, g_e, b_e):
                    def cb():
                        c = colorchooser.askcolor(color="#%02x%02x%02x" % self.group_colors[gg],
                                                  title=tr("ttf_pick_color") % gg)
                        if c[1] is None:
                            return
                        r, g_, b = (int(x) for x in c[0])
                        self.group_colors[gg] = (r, g_, b)
                        btn.configure(bg="#%02x%02x%02x" % (r, g_, b))
                        r_e.delete(0, "end"); r_e.insert(0, str(r))
                        g_e.delete(0, "end"); g_e.insert(0, str(g_))
                        b_e.delete(0, "end"); b_e.insert(0, str(b))
                    return cb

                def _make_rgb_apply(gg, btn, r_e, g_e, b_e):
                    def fn():
                        try:
                            r = max(0, min(255, int(float(r_e.get()))))
                            g_ = max(0, min(255, int(float(g_e.get()))))
                            b = max(0, min(255, int(float(b_e.get()))))
                        except (ValueError, TypeError):
                            return
                        self.group_colors[gg] = (r, g_, b)
                        btn.configure(bg="#%02x%02x%02x" % (r, g_, b))
                    return fn

                rgb_frame = ttk.Frame(rowf); rgb_frame.pack(side="left", padx=2)
                gw["rgb_vars"] = []
                _rgb_entries = []
                for idx, ch in enumerate(["R", "G", "B"]):
                    ttk.Label(rgb_frame, text=ch, font=("Consolas", 8, "bold"),
                             foreground="#666").pack(side="left")
                    v = tk.StringVar(value=str(col[idx]))
                    gw["rgb_vars"].append(v)
                    e = ttk.Entry(rgb_frame, textvariable=v, width=4)
                    e.pack(side="left", padx=1)
                    _rgb_entries.append(e)

                sw.configure(_make_cs(g, sw, _rgb_entries[0], _rgb_entries[1], _rgb_entries[2]))
                _apply_fn = _make_rgb_apply(g, sw, _rgb_entries[0], _rgb_entries[1], _rgb_entries[2])
                for e in _rgb_entries:
                    e.bind("<Return>", lambda ev, fn=_apply_fn: fn())
                    e.bind("<FocusOut>", lambda ev, fn=_apply_fn: fn())

                opa_frame = ttk.Frame(rowf); opa_frame.pack(side="left", padx=2)
                ttk.Label(opa_frame, text=tr("ttf_opacity"), font=("Microsoft YaHei", 8)).pack(side="left")
                opa_var = tk.IntVar(value=int(self.group_opacities[g]))
                gw["opacity_var"] = opa_var
                ttk.Entry(opa_frame, textvariable=opa_var, width=4).pack(side="left", padx=1)
                ttk.Label(opa_frame, text="%", font=("Microsoft YaHei", 8)).pack(side="left")

                gs = self.group_styles[g]
                gw["bold_var"] = tk.BooleanVar(value=gs["bold"])
                gw["italic_var"] = tk.BooleanVar(value=gs["italic"])
                ttk.Checkbutton(rowf, text=tr("ttf_bold"), variable=gw["bold_var"], width=5).pack(side="left", padx=2)
                ttk.Checkbutton(rowf, text=tr("ttf_italic"), variable=gw["italic_var"], width=5).pack(side="left", padx=2)

        ttk.Button(btn_f, text=tr("ttf_reset"), command=on_reset).pack(side="left", padx=8)

        def on_ok():
            # 收集每组的样式
            for g in self.groups:
                gw = self._dlg_group_widgets.get(g)
                if not gw:
                    continue
                self.group_styles[g]["bold"] = gw["bold_var"].get()
                self.group_styles[g]["italic"] = gw["italic_var"].get()
                # 从 RGB 输入框读取颜色值
                try:
                    r = max(0, min(255, int(float(gw["rgb_vars"][0].get()))))
                    g_ = max(0, min(255, int(float(gw["rgb_vars"][1].get()))))
                    b = max(0, min(255, int(float(gw["rgb_vars"][2].get()))))
                    self.group_colors[g] = (r, g_, b)
                except (ValueError, TypeError, KeyError):
                    pass
                try:
                    opa = int(float(gw["opacity_var"].get()))
                    self.group_opacities[g] = max(0, min(255, opa))
                except Exception:
                    self.group_opacities[g] = 255
            ttf = self._dlg_ttf_path.get().strip()
            out = self._dlg_out_dir.get().strip()
            # 相对路径解析为基于程序根目录的绝对路径
            if not os.path.isabs(ttf):
                ttf = os.path.join(app_base(), ttf)
            if not os.path.isabs(out):
                out = os.path.join(app_base(), out)
            if not ttf or not os.path.isfile(ttf):
                messagebox.showerror(tr("ttf_err_font"), tr("ttf_err_font_msg"))
                return
            if not out:
                messagebox.showerror(tr("ttf_err_out"), tr("ttf_err_out_msg"))
                return
            win.destroy()
            self._ttf_generate(ttf, out)

        ttk.Button(btn_f, text=tr("ttf_gen"), command=on_ok).pack(side="left", padx=16)

        ttk.Label(win, text=tr("ttf_tip"),
                  font=("Microsoft YaHei", 8), foreground="#888").pack()

    def _dlg_pick_ttf(self):
        p = filedialog.askopenfilename(
            title=tr("ttf_pick_font"), filetypes=[("Font (TTF/OTF)", "*.ttf;*.otf"), ("TTF", "*.ttf"), ("OTF", "*.otf")], initialdir=app_root())
        if p:
            self._dlg_ttf_path.set(p)

    def _dlg_pick_outdir(self):
        d = filedialog.askdirectory(
            title=tr("ttf_pick_out"), initialdir=app_root())
        if d:
            self._dlg_out_dir.set(d)

    def _ttf_generate(self, ttf_path, out_dir):
        """收集每组样式并启动 TTF 生成线程。"""
        gs_for_core = {}
        for g, st in self.group_styles.items():
            gs_for_core[g] = {
                "bold": st["bold"],
                "italic": st["italic"],
                "opacity": self.group_opacities.get(g, 255.0),
            }
        threading.Thread(target=self._ttf_worker,
                         args=(ttf_path, out_dir, gs_for_core),
                         daemon=True).start()

    def _ttf_worker(self, ttf, out_dir, group_styles):
        try:
            self.log(tr("ttf_gen_log") % os.path.basename(ttf))
            n = core.generate_from_ttf(ttf, out_dir, self.sprites, log=self.log,
                                       group_colors=self.group_colors,
                                       group_styles=group_styles)
            # 自动设为替换并加载预览
            for sp in self.sprites:
                if not sp.get("char"):
                    continue
                p = os.path.join(out_dir, sp["name"] + ".png")
                if os.path.exists(p):
                    st = self.row_by_name[sp["name"]]
                    st["repl"] = Image.open(p).convert("RGBA")
                    st["action"] = "replace"
                    st["ratio_var"].set(100)
                    st["ratio"] = 100.0
                    st["stretch"] = False
                    if "stretch_var" in st:
                        st["stretch_var"].set(False)
                    for k in st["action_states"]:
                        st["action_states"][k].set(k == "replace")
                    self.root.after(0, lambda s=st: self._refresh_preview(s))
            self.log(tr("ttf_done_log") % n)
            self.root.after(0, lambda: messagebox.showinfo(
                tr("ttf_done_title"), tr("ttf_done_msg") % (n, out_dir)))
        except Exception as e:
            self.log(tr("ttf_fail_log") % e)
            self.root.after(0, lambda ee=e: messagebox.showerror(tr("ttf_fail_title"), str(ee)))

    def on_build(self):
        out = self.out_dir.get()
        if not out:
            messagebox.showerror(tr("build_out_missing"), tr("build_out_missing_msg"))
            return
        # 相对路径解析为基于程序根目录的绝对路径
        if not os.path.isabs(out):
            out = os.path.join(app_base(), out)
        # 收集状态
        actions, replacements, ratios, stretches = {}, {}, {}, {}
        for st in self.rows:
            actions[st["name"]] = st["action"]
            ratios[st["name"]] = st["ratio"]
            stretches[st["name"]] = st.get("stretch", False)
            if st["action"] == "replace" and st["repl"] is not None:
                replacements[st["name"]] = st["repl"]
        n_rep = sum(1 for s in actions.values() if s == "replace")
        n_block = sum(1 for s in actions.values() if s == "block")
        if n_rep == 0 and n_block == 0:
            if not messagebox.askyesno(tr("confirm_title"), tr("confirm_msg")):
                return
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        mod_id = "gbfr.damagefont.editor.%s" % ts
        self.log(tr("build_start") % (n_rep, n_block, len(actions) - n_rep - n_block))
        threading.Thread(target=self._build_worker, args=(out, mod_id, actions, replacements, ratios, stretches), daemon=True).start()

    def _build_worker(self, out, mod_id, actions, replacements, ratios, stretches):
        try:
            zp = core.build_mod(actions, replacements, ratios, out,
                                 mod_id=mod_id, mod_name=self.mod_name.get(), log=self.log,
                                 stretches=stretches, sizes=self.dmg_sizes, colors=self.dmg_colors, range_alpha=self.dmg_range_alpha)
            self.root.after(0, lambda: messagebox.showinfo(tr("build_done_title"), tr("build_done_msg") % zp))
            self.root.after(0, lambda: os.startfile(os.path.dirname(zp)) if sys.platform == "win32" else None)
        except Exception as e:
            self.log(tr("build_fail_log") % e)
            self.root.after(0, lambda: messagebox.showerror(tr("build_fail_title"), str(e)))

    def on_export(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".json", filetypes=[("JSON", "*.json")], initialdir=app_root())
        if not p:
            return
        data = {"mod_name": self.mod_name.get(), "sprites": [],
                "dmg_sizes": dict(self.dmg_sizes)}
        for st in self.rows:
            entry = {"name": st["name"], "action": st["action"], "ratio": st["ratio"], "stretch": st.get("stretch", False)}
            if st["repl"] is not None:
                import io, base64
                buf = io.BytesIO()
                st["repl"].save(buf, format="PNG")
                entry["png_b64"] = base64.b64encode(buf.getvalue()).decode("ascii")
            data["sprites"].append(entry)
        json.dump(data, open(p, "w", encoding="utf-8"), ensure_ascii=False)
        self.log(tr("log_export") % p)

    def _embed_image(self, ws, pil, anchor):
        """把 PIL 图缩略后嵌入 Excel 单元格。"""
        thumb = pil.copy()
        thumb.thumbnail((64, 64))
        buf = io.BytesIO()
        thumb.save(buf, format="PNG")
        buf.seek(0)
        img = XLImage(buf)
        ws.add_image(img, anchor)

    def _exp_default_name(self, kind):
        # 按当前语言给出默认文件名(不含扩展名)，中英文不同
        zh = {"names": "元素名列表", "coords": "坐标表", "all": "全列导出", "pngdir": "原生图"}
        en = {"names": "element_names", "coords": "coordinates", "all": "all_columns", "pngdir": "native_pngs"}
        return (zh if LANG == "zh" else en).get(kind, "export")

    def on_export_col(self, c):
        if c == COL_NAME:
            p = filedialog.asksaveasfilename(
                title=tr("exp_names_title"),
                initialfile=self._exp_default_name("names"),
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")], initialdir=app_root())
            if not p:
                return
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "元素名"
            ws.append(["序号", "元素名"])
            for i, st in enumerate(self.rows, 1):
                ws.append([i, st["name"]])
            ws.freeze_panes = "A2"
            wb.save(p)
            self.log(tr("log_export_names") % p)
            messagebox.showinfo(tr("exp_names_title"), tr("exp_done") % p)
        elif c == COL_COORD:
            p = filedialog.asksaveasfilename(
                title=tr("exp_coords_title"),
                initialfile=self._exp_default_name("coords"),
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")], initialdir=app_root())
            if not p:
                return
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "坐标"
            ws.append(["序号", "元素名", "X", "Y", "宽", "高"])
            for i, st in enumerate(self.rows, 1):
                sp = st["sp"]
                ws.append([i, st["name"], sp["x"], sp["y"], sp["w"], sp["h"]])
            ws.freeze_panes = "A2"
            wb.save(p)
            self.log(tr("log_export_coords") % p)
            messagebox.showinfo(tr("exp_coords_title"), tr("exp_done") % p)
        elif c == COL_NATIVE:
            d = filedialog.askdirectory(title=tr("exp_pngs_title"),
                                        initialdir=os.path.join(app_root(), self._exp_default_name("pngdir")))
            if not d:
                return
            n = 0
            for st in self.rows:
                try:
                    st["native_pil"].save(os.path.join(d, st["name"] + ".png"), "PNG")
                    n += 1
                except Exception as e:
                    self.log(tr("png_export_fail", st["name"], e))
            self.log(tr("log_export_pngs") % (n, d))
            messagebox.showinfo(tr("exp_pngs_title"), tr("exp_pngs_done") % (n, d))

    def on_export_all(self):
        p = filedialog.asksaveasfilename(
            title=tr("exp_all_title"),
            initialfile=self._exp_default_name("all"),
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")], initialdir=app_root())
        if not p:
            return
        from openpyxl.styles import Font
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "全列导出"
        headers = ["元素名", "坐标(X,Y)", "尺寸(WxH)", "原生图", "动作", "替换图"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for st in self.rows:
            sp = st["sp"]
            ws.append([
                st["name"],
                "(%d,%d)" % (sp["x"], sp["y"]),
                "%dx%d" % (sp["w"], sp["h"]),
                "", st["action"], "",
            ])
            r = ws.max_row
            ws.row_dimensions[r].height = 52
            self._embed_image(ws, st["native_pil"], "D%d" % r)
            if st["repl"] is not None:
                self._embed_image(ws, st["repl"], "F%d" % r)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 11
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 11
        ws.freeze_panes = "A2"
        wb.save(p)
        self.log(tr("log_export_all") % p)
        messagebox.showinfo(tr("exp_all_title"), tr("exp_done") % p)

    def on_import(self):
        p = filedialog.askopenfilename(
            filetypes=[("JSON", "*.json")], initialdir=app_root())
        if not p:
            return
        data = json.load(open(p, encoding="utf-8"))
        self.mod_name.set(data.get("mod_name", self.mod_name.get()))
        for k, v in (data.get("dmg_sizes") or {}).items():
            try:
                self.dmg_sizes[k] = float(v)
            except Exception:
                pass
        m = {e["name"]: e for e in data.get("sprites", [])}
        import base64, io
        for st in self.rows:
            e = m.get(st["name"])
            if not e:
                continue
            st["action"] = e.get("action", "keep")
            st["ratio"] = e.get("ratio", 100.0)
            st["ratio_var"].set(st["ratio"])
            st["stretch"] = e.get("stretch", False)
            if "stretch_var" in st:
                st["stretch_var"].set(st["stretch"])
            for k in st["action_states"]:
                st["action_states"][k].set(k == st["action"])
            b64 = e.get("png_b64")
            if b64:
                try:
                    st["repl"] = Image.open(io.BytesIO(base64.b64decode(b64))).convert("RGBA")
                except Exception:
                    st["repl"] = None
            else:
                st["repl"] = None
            self._refresh_preview(st)
        self.log(tr("log_import_cfg") % p)

    def on_about(self):
        messagebox.showinfo(tr("about_title"), tr("about_body"))

    def on_tutorial(self):
        data = TUTORIAL_TEXT.get(LANG, TUTORIAL_TEXT["zh"])
        win = tk.Toplevel(self.root)
        win.title(data["title"])
        win.geometry("780x580")
        try:
            win.transient(self.root); win.grab_set()
        except Exception:
            pass
        sb = ttk.Scrollbar(win)
        sb.pack(side="right", fill="y")
        txt = tk.Text(win, wrap="word", padx=10, pady=10, yscrollcommand=sb.set,
                       font=("Microsoft YaHei", 10))
        txt.pack(side="left", fill="both", expand=True)
        sb.config(command=txt.yview)
        txt.insert("1.0", data["body"])
        txt.configure(state="disabled")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
